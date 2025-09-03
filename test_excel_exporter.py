"""
Тестовий файл для перевірки Excel експортера
"""
import os
import sys
from decimal import Decimal
from datetime import datetime

# Додаємо поточну директорію до шляху
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from models import Product, ParsingResult
from excel_exporter import ExcelExporter

def test_excel_exporter():
    """Тестує Excel експортер"""
    try:
        # Створюємо тестові дані
        products = [
            Product(
                name="Тестовий товар 1",
                price=Decimal("100.50"),
                product_url="https://www.olx.ua/test1",
                availability=True,
                sku="TEST001"
            ),
            Product(
                name="Тестовий товар 2",
                price=Decimal("250.00"),
                product_url="https://www.olx.ua/test2",
                availability=False,
                sku="TEST002"
            ),
            Product(
                name="Тестовий товар 3",
                price=Decimal("75.25"),
                product_url="https://www.olx.ua/test3",
                availability=True,
                sku="TEST003"
            )
        ]
        
        # Створюємо результат парсингу
        result = ParsingResult(
            success=True,
            products=products,
            total_products=len(products),
            parsing_time=1.5
        )
        
        # Тестуємо експортер
        print("🧪 Тестування Excel експортера...")
        
        exporter = ExcelExporter()
        
        # Експортуємо в Excel
        output_dir = "parsed_data"
        category_name = "Тестова категорія"
        
        print("\n📊 Перший експорт (створення нового файлу)...")
        filepath = exporter.export_to_excel(result, output_dir, category_name)
        
        if filepath and os.path.exists(filepath):
            print(f"✅ Excel файл успішно створено: {filepath}")
            print(f"📊 Розмір файлу: {os.path.getsize(filepath)} байт")
            
            # Перевіряємо чи файл можна відкрити
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                ws = wb.active
                print(f"📋 Назва листа: {ws.title}")
                print(f"📊 Кількість рядків: {ws.max_row}")
                print(f"📊 Кількість колонок: {ws.max_column}")
                wb.close()
                print("✅ Файл успішно відкрито та прочитано")
            except Exception as e:
                print(f"❌ Помилка при відкритті файлу: {e}")
        else:
            print("❌ Помилка при створенні Excel файлу")
            return
        
        # Тестуємо додавання дублікатів
        print("\n🔄 Тестування додавання дублікатів...")
        
        # Створюємо нові товари з дублікатами
        duplicate_products = [
            Product(
                name="Тестовий товар 1",  # Дублікат
                price=Decimal("110.00"),
                product_url="https://www.olx.ua/test1_duplicate",
                availability=True,
                sku="TEST001_DUP"
            ),
            Product(
                name="Новий товар 4",  # Новий товар
                price=Decimal("300.00"),
                product_url="https://www.olx.ua/test4",
                availability=True,
                sku="TEST004"
            ),
            Product(
                name="Тестовий товар 2",  # Дублікат
                price=Decimal("260.00"),
                product_url="https://www.olx.ua/test2_duplicate",
                availability=False,
                sku="TEST002_DUP"
            )
        ]
        
        duplicate_result = ParsingResult(
            success=True,
            products=duplicate_products,
            total_products=len(duplicate_products),
            parsing_time=0.5
        )
        
        # Експортуємо з дублікатами
        print("📊 Другий експорт (додавання до існуючого файлу)...")
        updated_filepath = exporter.export_to_excel(duplicate_result, output_dir, category_name)
        
        if updated_filepath and os.path.exists(updated_filepath):
            print(f"✅ Файл успішно оновлено: {updated_filepath}")
            print(f"📊 Новий розмір файлу: {os.path.getsize(updated_filepath)} байт")
            
            # Перевіряємо оновлений файл
            try:
                from openpyxl import load_workbook
                wb = load_workbook(updated_filepath)
                ws = wb.active
                print(f"📋 Назва листа: {ws.title}")
                print(f"📊 Кількість рядків: {ws.max_row}")
                print(f"📊 Кількість колонок: {ws.max_column}")
                wb.close()
                print("✅ Оновлений файл успішно відкрито та прочитано")
            except Exception as e:
                print(f"❌ Помилка при відкритті оновленого файлу: {e}")
        else:
            print("❌ Помилка при оновленні Excel файлу")
            
    except Exception as e:
        print(f"❌ Помилка при тестуванні: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_exporter()
