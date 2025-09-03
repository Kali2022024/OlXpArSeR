"""
Модуль для експорту даних в Excel формат
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


from models import Product, ParsingResult

class ExcelExporter:
    """Клас для експорту даних в Excel формат"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.existing_products: Set[str] = set()  # Множина існуючих назв товарів
    
    def create_workbook(self, category_name: str) -> str:
        """Створює новий Excel файл з назвою категорії"""
        try:
            # Створюємо робочу книгу
            self.workbook = Workbook()
            
            # Видаляємо стандартний лист
            self.workbook.remove(self.workbook.active)
            
            # Створюємо новий лист з назвою категорії
            sheet_name = self._sanitize_sheet_name(category_name)
            self.worksheet = self.workbook.create_sheet(title=sheet_name)
            
            return sheet_name
            
        except Exception as e:
            raise Exception(f"Помилка при створенні робочої книги: {e}")
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Очищає назву для використання як назва листа Excel"""
        # Excel має обмеження на назви листів
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Обмежуємо довжину назви
        if len(sanitized) > 31:
            sanitized = sanitized[:31]
        
        return sanitized
    
    def find_existing_excel_file(self, output_directory: str, category_name: str) -> Optional[str]:
        """Знаходить існуючий Excel файл для категорії"""
        try:
            if not os.path.exists(output_directory):
                return None
            
            # Шукаємо файли з назвою категорії
            for filename in os.listdir(output_directory):
                if filename.startswith(f"Олх_{category_name}_") and filename.endswith('.xlsx'):
                    filepath = os.path.join(output_directory, filename)
                    if os.path.isfile(filepath):
                        return filepath
            
            return None
            
        except Exception as e:
            print(f"Попередження: Помилка при пошуку існуючого файлу: {e}")
            return None
    
    def load_existing_workbook(self, filepath: str, category_name: str) -> bool:
        """Завантажує існуючий Excel файл"""
        try:
            self.workbook = load_workbook(filepath)
            
            # Знаходимо лист з назвою категорії
            sheet_name = self._sanitize_sheet_name(category_name)
            if sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[sheet_name]
            else:
                # Створюємо новий лист якщо не знайдено
                self.worksheet = self.workbook.create_sheet(title=sheet_name)
                self.setup_headers()
                return True
            
            # Завантажуємо існуючі назви товарів для перевірки дублікатів
            self._load_existing_products()
            
            return True
            
        except Exception as e:
            print(f"Попередження: Помилка при завантаженні існуючого файлу: {e}")
            return False
    
    def _load_existing_products(self):
        """Завантажує існуючі назви товарів з листа"""
        try:
            self.existing_products.clear()
            
            # Перевіряємо чи є заголовки
            if self.worksheet.max_row < 2:
                return
            
            # Читаємо назви товарів з першої колонки (починаючи з 2-го рядка)
            for row in range(2, self.worksheet.max_row + 1):
                cell_value = self.worksheet.cell(row=row, column=1).value
                if cell_value:
                    # Нормалізуємо назву для порівняння
                    normalized_name = self._normalize_product_name(str(cell_value))
                    self.existing_products.add(normalized_name)
            
            print(f"📋 Завантажено {len(self.existing_products)} існуючих товарів")
            
        except Exception as e:
            print(f"Попередження: Помилка при завантаженні існуючих товарів: {e}")
    
    def _normalize_product_name(self, name: str) -> str:
        """Нормалізує назву товару для порівняння"""
        # Приводимо до нижнього регістру та видаляємо зайві пробіли
        normalized = name.lower().strip()
        
        # Видаляємо зайві пробіли між словами
        normalized = ' '.join(normalized.split())
        
        return normalized
    
    def is_duplicate_product(self, product: Product) -> bool:
        """Перевіряє чи є товар дублікатом"""
        normalized_name = self._normalize_product_name(product.name)
        return normalized_name in self.existing_products
    
    def setup_headers(self):
        """Налаштовує заголовки колонок"""
        try:
            headers = [
                "Назва товару",
                "Ціна",
                "Наявність",
                "Посилання"
            ]
            
            # Додаємо заголовки
            for col, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=1, column=col, value=header)
                
                # Стилізуємо заголовки
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоматично налаштовуємо ширину колонок
            for col in range(1, len(headers) + 1):
                try:
                    column_letter = get_column_letter(col)
                    if col == 1:  # Назва товару
                        self.worksheet.column_dimensions[column_letter].width = 80
                    elif col == 2:  # Ціна
                        self.worksheet.column_dimensions[column_letter].width = 15
                    elif col == 3:  # Наявність
                        self.worksheet.column_dimensions[column_letter].width = 20
                    elif col == 4:  # Посилання
                        self.worksheet.column_dimensions[column_letter].width = 60
                except Exception as e:
                    print(f"Попередження: Помилка при налаштуванні ширини колонки {col}: {e}")
                    continue
                    
        except Exception as e:
            raise Exception(f"Помилка при налаштуванні заголовків: {e}")
    
    def add_product_data(self, products: List[Product]):
        """Додає дані про товари в Excel з перевіркою дублікатів"""
        new_products_count = 0
        duplicate_products_count = 0
        
        for product in products:
            try:
                # Перевіряємо чи є дублікат
                if self.is_duplicate_product(product):
                    print(f"⚠️  Пропущено дублікат: {product.name}")
                    duplicate_products_count += 1
                    continue
                
                # Знаходимо наступний вільний рядок
                next_row = self.worksheet.max_row + 1
                
                # Назва товару
                name_cell = self.worksheet.cell(row=next_row, column=1, value=product.name or "Без назви")
                
                # Ціна
                price_value = str(product.price) if product.price else "0"
                price_cell = self.worksheet.cell(row=next_row, column=2, value=price_value)
                price_cell.alignment = Alignment(horizontal="right")
                
                # Наявність
                availability_text = "✅ В наявності" if product.availability else "❌ Немає в наявності"
                availability_cell = self.worksheet.cell(row=next_row, column=3, value=availability_text)
                availability_cell.alignment = Alignment(horizontal="center")
                
                # Посилання (гіперпосилання)
                link_cell = self.worksheet.cell(row=next_row, column=4, value="Перейти до товару")
                link_cell.font = Font(color="0000FF", underline="single")
                link_cell.alignment = Alignment(horizontal="center")
                
                # Додаємо гіперпосилання
                if product.product_url:
                    # Створюємо гіперпосилання безпосередньо в комірці
                    link_cell.hyperlink = product.product_url
                    link_cell.value = "Перейти до товару"
                
                # Додаємо назву до множини існуючих товарів
                normalized_name = self._normalize_product_name(product.name)
                self.existing_products.add(normalized_name)
                
                new_products_count += 1
                    
            except Exception as e:
                print(f"Попередження: Помилка при додаванні товару: {e}")
                continue
        
        print(f"📊 Додано нових товарів: {new_products_count}")
        print(f"🔄 Пропущено дублікатів: {duplicate_products_count}")
        
        return new_products_count, duplicate_products_count
    
    def apply_alternating_row_colors(self):
        """Застосовує чергування кольорів рядків для кращої читабельності"""
        try:
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            max_row = self.worksheet.max_row
            max_col = self.worksheet.max_column
            
            if max_row > 1 and max_col > 0:
                for row in range(2, max_row + 1):
                    if row % 2 == 0:  # Парні рядки
                        for col in range(1, max_col + 1):
                            try:
                                cell = self.worksheet.cell(row=row, column=col)
                                cell.fill = light_fill
                            except Exception as e:
                                print(f"Попередження: Помилка при стилізації комірки {row}:{col}: {e}")
                                continue
        except Exception as e:
            print(f"Попередження: Помилка при застосуванні кольорів: {e}")
    
    def save_workbook(self, output_directory: str, category_name: str) -> str:
        """Зберігає Excel файл"""
        try:
            # Створюємо директорію якщо не існує
            os.makedirs(output_directory, exist_ok=True)
            
            # Формуємо назву файлу
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Олх_{category_name}_{timestamp}.xlsx"
            filepath = os.path.join(output_directory, filename)
            
            # Зберігаємо файл
            self.workbook.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Помилка при збереженні файлу: {e}")
    
    def export_to_excel(self, result: ParsingResult, output_directory: str, category_name: str) -> str:
        """Основна функція експорту в Excel"""
        try:
            # Перевіряємо наявність даних
            if not result.products:
                raise Exception("Немає даних для експорту")
            
            # Спочатку шукаємо існуючий файл
            existing_file = self.find_existing_excel_file(output_directory, category_name)
            
            if existing_file:
                print(f"📁 Знайдено існуючий файл: {existing_file}")
                print("🔄 Додаємо нові дані до існуючого файлу...")
                
                # Завантажуємо існуючий файл
                if self.load_existing_workbook(existing_file, category_name):
                    # Додаємо нові дані
                    self.add_product_data(result.products)
                    
                    # Застосовуємо стилі
                    self.apply_alternating_row_colors()
                    
                    # Зберігаємо оновлений файл
                    self.workbook.save(existing_file)
                    return existing_file
                else:
                    print("⚠️  Не вдалося завантажити існуючий файл, створюємо новий")
            
            # Створюємо нову робочу книгу
            self.create_workbook(category_name)
            
            # Налаштовуємо заголовки
            self.setup_headers()
            
            # Додаємо дані про товари
            self.add_product_data(result.products)
            
            # Застосовуємо стилі
            self.apply_alternating_row_colors()
            
            # Зберігаємо файл
            filepath = self.save_workbook(output_directory, category_name)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Помилка при експорті в Excel: {e}")
        finally:
            # Закриваємо робочу книгу
            if self.workbook:
                self.workbook.close()
